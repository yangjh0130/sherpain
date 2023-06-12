from tenable.io import TenableIO
from openpyxl import load_workbook
import time 

########################################################
#                                                      #
#                                                      #
# asset_file_path, tag_file_path, tio 정보 수정 후 실행 #
#                                                      #
#                                                      #
########################################################


# Asset 엑셀 파일 경로
asset_file_path = "C:/Users/YJH/Desktop/python/tenable/asset_tag/AddAsset.xlsx"

# Tag 엑셀 파일 경로
tag_file_path = "C:/Users/YJH/Desktop/python/tenable/asset_tag/CreateTag.xlsx"

# 1. 접속 정보 입력 : TenableIO('ACCESS_KEY', 'SECRET_KEY')
tio = TenableIO('3210fd3958dd5243fa4580dc5e088fd4bf2a6739896ff28b1de4ea567ed3dd54', 'cd95b22df1da6d80c5e5b4ce52a257f279ecf7825f45aa7f6c313863623c98a2')

# asset 등록
def reg_asset():

    # 엑셀파일 읽어오기
    load_wd = load_workbook(asset_file_path, data_only=True)
    load_ws=load_wd['Sheet1']

    # 데이터 읽어 오기
    data = []
    for row in load_ws.rows:
        data.append([row[0].value, row[1].value])

    num = 0
    print("========================== Asset 등록을 시작합니다. ==========================")
                            
    for j in data:
        tio.assets.asset_import('custom', {'ipv4': j})

        print(  str(num+1) + "..         IP 주소 :"+ str(j) + "      등록 완료") 
        num+=1
    
    print("========================== Asset 등록이 완료되었습니다. ==========================")

'''
# tag 생성
def cre_tags():

    # 엑셀파일 읽어오기
    load_wd = load_workbook(tag_file_path, data_only=True)
    load_ws=load_wd['Sheet1']

    # 데이터 읽어 오기
    data = []
    for row in load_ws.rows:
        data.append([row[0].value, row[1].value,row[2].value])

    num=0
    print("========================== Tag 생성을 시작합니다. ==========================")

    for i in data:
        tio.tags.create(i[1], i[2])

        print(str(num+1) + "..         Category: " + i[1] + ", Value: " + i[2] + " 등록 완료")
        num+=1

    print("========================== Tag 생성이 완료되었습니다. ==========================")
'''

reg_asset()
time.sleep(1)
'''
cre_tags()'''
